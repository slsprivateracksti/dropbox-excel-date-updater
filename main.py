"""
main.py
Dropbox Excel 日付更新ツール v2
タブ構成:
  ① 単体実行
  ② 一括実行（F案: 左リスト＋右詳細）
  ③ 履歴・パターン管理
"""

import copy
import os
import re
import threading
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk

from config_loader import ConfigLoader, DEFAULT_JSON_PATH
from excel_updater import ExcelUpdater
from log_manager import LogManager
from pattern_store import PatternStore
from scanner import DropboxScanner
import cell_selector

try:
    from tkcalendar import Calendar
    _CAL_AVAILABLE = True
except ImportError:
    _CAL_AVAILABLE = False


# ══════════════════════════════════════════════════════
# ユーティリティ
# ══════════════════════════════════════════════════════

def get_dropbox_path() -> str:
    return os.path.join(os.environ.get("USERPROFILE", ""), "Dropbox")


def validate_dropbox_path(path: str) -> bool:
    return os.path.isdir(path)


def _col_to_idx(col: str) -> int:
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def _idx_to_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def expand_cell_range(token: str) -> list[str]:
    """
    "H3:H20" -> ["H3","H4",...,"H20"]
    "B2:D4"  -> ["B2","B3","B4","C2",...]  (矩形展開)
    "I2"     -> ["I2"]
    """
    token = token.strip().upper()
    m = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", token)
    if not m:
        return [token] if token else []
    col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    c1i, c2i = _col_to_idx(col1), _col_to_idx(col2)
    result = []
    for ci in range(min(c1i, c2i), max(c1i, c2i) + 1):
        for ri in range(min(row1, row2), max(row1, row2) + 1):
            result.append(f"{_idx_to_col(ci)}{ri}")
    return result


def parse_cells(text: str) -> list[str]:
    """カンマ区切り・範囲表記混在のセル文字列をフラットなリストに展開。"""
    result = []
    for token in text.split(","):
        result.extend(expand_cell_range(token))
    return result


def compress_cells(cells: list[str]) -> str:
    """連続する同列セルを範囲表記に圧縮: ["H3","H4","H5"] -> "H3:H5" """
    if not cells:
        return ""

    def col_row(c: str):
        m = re.fullmatch(r"([A-Z]+)(\d+)", c.strip().upper())
        return (m.group(1), int(m.group(2))) if m else (c.upper(), 0)

    parsed = [col_row(c) for c in cells]
    groups: list[list] = [[parsed[0]]]
    for col, row in parsed[1:]:
        pc, pr = groups[-1][-1]
        if col == pc and row == pr + 1:
            groups[-1].append((col, row))
        else:
            groups.append([(col, row)])

    tokens = []
    for g in groups:
        if len(g) == 1:
            tokens.append(f"{g[0][0]}{g[0][1]}")
        else:
            tokens.append(f"{g[0][0]}{g[0][1]}:{g[-1][0]}{g[-1][1]}")
    return ", ".join(tokens)


def cells_to_str(cells: list[str]) -> str:
    return compress_cells(cells)


def format_date_preview(fmt: str, d: date | None = None) -> str:
    if d is None:
        d = date.today()
    result = fmt
    result = result.replace("YYYY", f"{d.year:04d}")
    result = result.replace("MM",   f"{d.month:02d}")
    result = result.replace("M",    str(d.month))
    result = result.replace("DD",   f"{d.day:02d}")
    result = result.replace("D",    str(d.day))
    return result


def search_rows(rows: list[dict], keys: list[str], keyword: str, mode: str) -> list[dict]:
    words = [w.strip().lower() for w in keyword.split() if w.strip()]
    if not words:
        return rows
    result = []
    for row in rows:
        haystack = " ".join(str(row.get(k, "")) for k in keys).lower()
        hit = all(w in haystack for w in words) if mode == "AND" else any(w in haystack for w in words)
        if hit:
            result.append(row)
    return result


def center_window(win: tk.Toplevel, parent: tk.Widget):
    win.update_idletasks()
    pw = parent.winfo_rootx() + parent.winfo_width()  // 2
    ph = parent.winfo_rooty() + parent.winfo_height() // 2
    win.geometry(f"+{pw - win.winfo_width()//2}+{ph - win.winfo_height()//2}")


# ══════════════════════════════════════════════════════
# カレンダーポップアップ
# ══════════════════════════════════════════════════════

class CalendarPopup(tk.Toplevel):
    def __init__(self, parent, date_var: tk.StringVar):
        super().__init__(parent)
        self.title("日付を選択")
        self.resizable(False, False)
        self.grab_set()
        self.date_var = date_var
        try:
            initial = date.fromisoformat(date_var.get())
        except ValueError:
            initial = date.today()
        self.cal = Calendar(
            self, selectmode="day",
            year=initial.year, month=initial.month, day=initial.day,
            date_pattern="yyyy-mm-dd", locale="ja_JP",
            showweeknumbers=False, firstweekday="sunday",
        )
        self.cal.pack(padx=12, pady=(12, 6))
        ttk.Button(self, text="選択", command=self._on_select).pack(pady=(0, 12))
        center_window(self, parent)

    def _on_select(self):
        self.date_var.set(self.cal.get_date())
        self.destroy()


def pick_date(parent, date_var):
    if _CAL_AVAILABLE:
        CalendarPopup(parent, date_var)
    else:
        messagebox.showwarning("tkcalendar 未インストール",
                               "pip install tkcalendar でインストールしてください。", parent=parent)


# ══════════════════════════════════════════════════════
# 日付形式設定ダイアログ
# ══════════════════════════════════════════════════════

class DateFormatDialog(tk.Toplevel):
    PRESETS = ["M/D", "YYYY/MM/DD", "YYYY-MM-DD", "YYYY年MM月DD日",
               "MM/DD/YYYY", "M月D日", "DD/MM/YYYY"]

    def __init__(self, parent, current_fmt: str, callback):
        super().__init__(parent)
        self.title("日付形式の設定")
        self.resizable(False, False)
        self.grab_set()
        self._callback = callback
        self._fmt_var  = tk.StringVar(value=current_fmt)

        f = ttk.Frame(self, padding=16)
        f.pack(fill="both", expand=True)

        ttk.Label(f, text="形式：").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(f, textvariable=self._fmt_var, width=24,
                  font=("Yu Gothic UI", 11)).grid(row=0, column=1, columnspan=2, sticky="ew", pady=4)
        self._fmt_var.trace_add("write", lambda *_: self._update_preview())

        self._preview_var = tk.StringVar()
        ttk.Label(f, text="プレビュー：").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Label(f, textvariable=self._preview_var,
                  font=("Yu Gothic UI", 13, "bold"), foreground="#0066cc").grid(
            row=1, column=1, columnspan=2, sticky="w", pady=4)

        ttk.Label(f, text="書式記号：\nYYYY=年  MM=月(0埋め)  M=月(0なし)\nDD=日(0埋め)  D=日(0なし)",
                  foreground="#555", justify="left").grid(
            row=2, column=0, columnspan=3, sticky="w", pady=(2, 8))

        ttk.Label(f, text="よく使う形式：").grid(row=3, column=0, columnspan=3, sticky="w")
        pf = ttk.Frame(f)
        pf.grid(row=4, column=0, columnspan=3, sticky="w", pady=4)
        for i, p in enumerate(self.PRESETS):
            ttk.Button(pf, text=p, width=13,
                       command=lambda v=p: self._fmt_var.set(v)).grid(
                row=i // 4, column=i % 4, padx=3, pady=2)

        ttk.Separator(f, orient="horizontal").grid(
            row=5, column=0, columnspan=3, sticky="ew", pady=8)
        bf = ttk.Frame(f)
        bf.grid(row=6, column=0, columnspan=3)
        ttk.Button(bf, text="OK",       command=self._ok,      width=10).pack(side="left", padx=6)
        ttk.Button(bf, text="キャンセル", command=self.destroy, width=10).pack(side="left", padx=6)

        self._update_preview()
        center_window(self, parent)

    def _update_preview(self):
        self._preview_var.set(format_date_preview(self._fmt_var.get()))

    def _ok(self):
        self._callback(self._fmt_var.get())
        self.destroy()


# ══════════════════════════════════════════════════════
# セル選択ダイアログ（Excel連携）
# ══════════════════════════════════════════════════════

class CellSelectDialog(tk.Toplevel):
    def __init__(self, parent, title: str, file_path: str, current: str, callback):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.grab_set()
        self._file_path = file_path
        self._callback  = callback

        ttk.Label(self, text=(
            "① 下の「Excelを開く」ボタンでファイルを開いてください。\n"
            "② Excel でセルを選択（複数可・範囲選択可）してください。\n"
            "③「選択中のセルを取得」を押すと下の欄に反映されます。\n"
            "④「確定」で閉じます。手入力での修正も可能です。\n"
            "　 例: I2, H3:H20, C5"
        ), justify="left", wraplength=420).pack(padx=16, pady=(14, 6))

        self._var = tk.StringVar(value=current)
        ttk.Entry(self, textvariable=self._var, width=50,
                  font=("Yu Gothic UI", 11)).pack(padx=16, pady=4)

        bf = ttk.Frame(self)
        bf.pack(padx=16, pady=6)
        ttk.Button(bf, text="Excelを開く",       command=self._open_excel).grid(row=0, column=0, padx=4)
        ttk.Button(bf, text="選択中のセルを取得", command=self._get_cells).grid(row=0, column=1, padx=4)
        ttk.Button(bf, text="確定",               command=self._confirm).grid(row=0, column=2, padx=4)
        ttk.Button(bf, text="キャンセル",          command=self.destroy).grid(row=0, column=3, padx=4)

        self._status = tk.StringVar(value="")
        ttk.Label(self, textvariable=self._status, foreground="#555").pack(pady=(0, 10))
        center_window(self, parent)

    def _open_excel(self):
        if not self._file_path:
            self._status.set("⚠ ファイルが指定されていません。"); return
        ok = cell_selector.open_file_in_excel(self._file_path)
        self._status.set("✓ Excel を開きました。" if ok else "⚠ Excel を開けませんでした。手動で開いてください。")

    def _get_cells(self):
        if not cell_selector.is_available():
            self._status.set("⚠ pywin32 未インストール（pip install pywin32）。手入力してください。"); return
        cells = cell_selector.get_selected_cells(self._file_path or None)
        if cells:
            self._var.set(compress_cells(cells))
            self._status.set(f"✓ {len(cells)} セル取得しました。")
        else:
            self._status.set("⚠ セルを取得できませんでした。Excel でセルを選択後に再試行してください。")

    def _confirm(self):
        self._callback(self._var.get())
        self.destroy()


# ══════════════════════════════════════════════════════
# 確認ダイアログ（3択）
# ══════════════════════════════════════════════════════

class ConfirmDialog(tk.Toplevel):
    RESULT_YES  = "yes"
    RESULT_NO   = "no"
    RESULT_OPEN = "open"

    def __init__(self, parent, message: str):
        super().__init__(parent)
        self.title("確認")
        self.resizable(False, False)
        self.grab_set()
        self.result = self.RESULT_NO
        ttk.Label(self, text=message, wraplength=360, justify="left").pack(padx=20, pady=(18, 10))
        bf = ttk.Frame(self)
        bf.pack(pady=(0, 16))
        ttk.Button(bf, text="はい",         command=self._yes,  width=12).grid(row=0, column=0, padx=6)
        ttk.Button(bf, text="いいえ",       command=self._no,   width=12).grid(row=0, column=1, padx=6)
        ttk.Button(bf, text="ファイルを開く", command=self._open, width=14).grid(row=0, column=2, padx=6)
        center_window(self, parent)
        self.wait_window()

    def _yes(self):  self.result = self.RESULT_YES;  self.destroy()
    def _no(self):   self.result = self.RESULT_NO;   self.destroy()
    def _open(self): self.result = self.RESULT_OPEN; self.destroy()


# ══════════════════════════════════════════════════════
# 保存済みパターン選択ダイアログ（②一括実行用）
# ══════════════════════════════════════════════════════

class PatternSelectDialog(tk.Toplevel):
    def __init__(self, parent, store: "PatternStore", callback):
        super().__init__(parent)
        self.title("パターンを選択して追加")
        self.resizable(True, False)
        self.grab_set()
        self._store    = store
        self._callback = callback

        ttk.Label(self, text="追加するパターンを選択してください（複数選択可）",
                  font=("Yu Gothic UI", 10)).pack(padx=12, pady=(12, 4))

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=12, pady=4)
        self._lb = tk.Listbox(frame, selectmode="extended", width=54, height=14,
                              font=("Yu Gothic UI", 10))
        sb = ttk.Scrollbar(frame, orient="vertical", command=self._lb.yview)
        self._lb.config(yscrollcommand=sb.set)
        self._lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self._patterns = store.get_all()
        for p in self._patterns:
            fname = p.get("facility_name", "")
            pname = p.get("pattern_name", "")
            self._lb.insert("end", f"{fname}  [{pname}]")

        bf = ttk.Frame(self)
        bf.pack(pady=8)
        ttk.Button(bf, text="追加",       command=self._add,    width=12).pack(side="left", padx=6)
        ttk.Button(bf, text="キャンセル", command=self.destroy, width=12).pack(side="left", padx=6)
        center_window(self, parent)

    def _add(self):
        sel = self._lb.curselection()
        if not sel:
            messagebox.showinfo("選択なし", "パターンを選択してください。", parent=self); return
        self._callback([self._patterns[i] for i in sel])
        self.destroy()


# ══════════════════════════════════════════════════════
# パターン編集ダイアログ
# ══════════════════════════════════════════════════════

class PatternEditDialog(tk.Toplevel):
    def __init__(self, parent, dropbox_root: str, pattern: dict | None,
                 callback, allow_rename_save: bool = False):
        super().__init__(parent)
        self.title("パターン設定")
        self.resizable(False, False)
        self.grab_set()
        self._dropbox_root = dropbox_root
        self._callback     = callback
        self._allow_rename = allow_rename_save
        self._sheet_names: list[str] = []

        p = pattern or {}
        self._pname_var   = tk.StringVar(value=p.get("pattern_name",  ""))
        self._fname_var   = tk.StringVar(value=p.get("facility_name", ""))
        self._fpath_var   = tk.StringVar(value=p.get("file_path",     ""))
        self._sheet_var   = tk.StringVar(value=p.get("sheet_name",    ""))
        self._target_var  = tk.StringVar(value=cells_to_str(p.get("target_cells",  [])))
        self._exclude_var = tk.StringVar(value=cells_to_str(p.get("exclude_cells", [])))
        self._dfmt_var    = tk.StringVar(value=p.get("date_format", "M/D"))

        self._build_ui()
        center_window(self, parent)
        if self._fpath_var.get():
            self._reload_sheets()

    def _build_ui(self):
        f = ttk.Frame(self, padding=16)
        f.pack(fill="both", expand=True)
        f.columnconfigure(1, weight=1)

        for i, (label, var) in enumerate([("パターン名：", self._pname_var),
                                           ("施設名：",     self._fname_var)]):
            ttk.Label(f, text=label).grid(row=i, column=0, sticky="w", padx=(0, 8), pady=4)
            ttk.Entry(f, textvariable=var, width=30).grid(row=i, column=1, columnspan=2, sticky="ew", pady=4)

        ttk.Label(f, text="ファイル：").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(f, textvariable=self._fpath_var, width=28).grid(row=2, column=1, sticky="ew", pady=4)
        ttk.Button(f, text="参照...", command=self._browse_file).grid(row=2, column=2, padx=(4, 0))

        ttk.Label(f, text="シート：").grid(row=3, column=0, sticky="w", pady=4)
        self._sheet_cb = ttk.Combobox(f, textvariable=self._sheet_var, values=[], state="readonly", width=26)
        self._sheet_cb.grid(row=3, column=1, columnspan=2, sticky="ew", pady=4)

        # 日付形式
        ttk.Label(f, text="日付形式：").grid(row=4, column=0, sticky="w", pady=4)
        dfmt_f = ttk.Frame(f)
        dfmt_f.grid(row=4, column=1, columnspan=2, sticky="w", pady=4)
        ttk.Entry(dfmt_f, textvariable=self._dfmt_var, width=14).pack(side="left")
        ttk.Button(dfmt_f, text="⚙設定", width=6,
                   command=lambda: DateFormatDialog(self, self._dfmt_var.get(),
                                                   lambda v: self._dfmt_var.set(v))).pack(side="left", padx=4)
        self._dfmt_preview = tk.StringVar()
        ttk.Label(dfmt_f, textvariable=self._dfmt_preview, foreground="#0066cc").pack(side="left", padx=4)
        self._dfmt_var.trace_add("write", lambda *_: self._dfmt_preview.set(
            f"→ {format_date_preview(self._dfmt_var.get())}"))
        self._dfmt_preview.set(f"→ {format_date_preview(self._dfmt_var.get())}")

        # 対象セル
        ttk.Label(f, text="対象セル：").grid(row=5, column=0, sticky="w", pady=4)
        tf = ttk.Frame(f)
        tf.grid(row=5, column=1, columnspan=2, sticky="ew", pady=4)
        ttk.Entry(tf, textvariable=self._target_var, width=20).pack(side="left", fill="x", expand=True)
        ttk.Button(tf, text="Excelで選択", command=self._select_target).pack(side="left", padx=3)
        ttk.Button(tf, text="✕", width=2, command=lambda: self._target_var.set("")).pack(side="left")

        # 除外セル
        ttk.Label(f, text="除外セル：").grid(row=6, column=0, sticky="w", pady=4)
        ef = ttk.Frame(f)
        ef.grid(row=6, column=1, columnspan=2, sticky="ew", pady=4)
        ttk.Entry(ef, textvariable=self._exclude_var, width=20).pack(side="left", fill="x", expand=True)
        ttk.Button(ef, text="Excelで選択", command=self._select_exclude).pack(side="left", padx=3)
        ttk.Button(ef, text="✕", width=2, command=lambda: self._exclude_var.set("")).pack(side="left")

        ttk.Separator(f, orient="horizontal").grid(row=7, column=0, columnspan=3, sticky="ew", pady=10)

        bf = ttk.Frame(f)
        bf.grid(row=8, column=0, columnspan=3)
        ttk.Button(bf, text="保存",       command=self._save,    width=10).pack(side="left", padx=4)
        if self._allow_rename:
            ttk.Button(bf, text="別名で保存", command=self._save_as, width=12).pack(side="left", padx=4)
        ttk.Button(bf, text="キャンセル", command=self.destroy,  width=10).pack(side="left", padx=4)

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択", initialdir=self._dropbox_root,
            filetypes=[("Excel ファイル", "*.xlsx *.xlsm *.xls")], parent=self)
        if path:
            self._fpath_var.set(os.path.relpath(path, self._dropbox_root))
            self._reload_sheets()

    def _reload_sheets(self):
        full = os.path.join(self._dropbox_root, self._fpath_var.get())
        try:
            import openpyxl
            wb = openpyxl.load_workbook(full, read_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
            self._sheet_cb["values"] = self._sheet_names
            if self._sheet_names and not self._sheet_var.get():
                self._sheet_cb.current(0)
        except Exception:
            self._sheet_names = []

    def _full_path(self) -> str:
        return os.path.join(self._dropbox_root, self._fpath_var.get())

    def _select_target(self):
        CellSelectDialog(self, "対象セルを選択", self._full_path(),
                         self._target_var.get(), lambda v: self._target_var.set(v))

    def _select_exclude(self):
        CellSelectDialog(self, "除外セルを選択", self._full_path(),
                         self._exclude_var.get(), lambda v: self._exclude_var.set(v))

    def _build_pattern(self, name: str) -> dict | None:
        if not name:
            messagebox.showwarning("入力エラー", "パターン名を入力してください。", parent=self); return None
        if not self._fpath_var.get().strip():
            messagebox.showwarning("入力エラー", "ファイルを選択してください。", parent=self); return None
        if not self._sheet_var.get().strip():
            messagebox.showwarning("入力エラー", "シートを選択してください。", parent=self); return None
        return {
            "pattern_name" : name,
            "facility_name": self._fname_var.get().strip(),
            "file_path"    : self._fpath_var.get().strip(),
            "sheet_name"   : self._sheet_var.get().strip(),
            "date_format"  : self._dfmt_var.get().strip(),
            "target_cells" : parse_cells(self._target_var.get()),
            "exclude_cells": parse_cells(self._exclude_var.get()),
            "last_executed": "",
            "last_status"  : "",
        }

    def _save(self):
        p = self._build_pattern(self._pname_var.get().strip())
        if p:
            self._callback(p)
            self.destroy()

    def _save_as(self):
        dlg = tk.Toplevel(self)
        dlg.title("別名で保存")
        dlg.resizable(False, False)
        dlg.grab_set()
        var = tk.StringVar(value=self._pname_var.get() + "_コピー")
        ttk.Label(dlg, text="新しいパターン名：").pack(padx=16, pady=(14, 4))
        ttk.Entry(dlg, textvariable=var, width=28).pack(padx=16)

        def do_save():
            p = self._build_pattern(var.get().strip())
            if p:
                self._callback(p)
                dlg.destroy()
                self.destroy()

        ttk.Button(dlg, text="保存", command=do_save, width=12).pack(pady=10)
        center_window(dlg, self)


# ══════════════════════════════════════════════════════
# タブ① 単体実行
# ══════════════════════════════════════════════════════

class SingleTab(ttk.Frame):
    def __init__(self, parent, app: "DateUpdateApp"):
        super().__init__(parent, padding=16)
        self._app = app
        self._sheet_names: list[str] = []
        self._build()

    def _build(self):
        self.columnconfigure(1, weight=1)

        # タイトル + 全てクリア
        title_row = ttk.Frame(self)
        title_row.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        ttk.Label(title_row, text="単体実行",
                  font=("Yu Gothic UI", 13, "bold")).pack(side="left")
        ttk.Button(title_row, text="全てクリア",
                   command=self._clear_all, width=10).pack(side="right")

        # 施設名
        ttk.Label(self, text="施設名：").grid(row=1, column=0, sticky="w", pady=5)
        fn_f = ttk.Frame(self)
        fn_f.grid(row=1, column=1, columnspan=2, sticky="ew", pady=5)
        self._fname_var = tk.StringVar()
        ttk.Entry(fn_f, textvariable=self._fname_var, width=30).pack(side="left", fill="x", expand=True)
        ttk.Button(fn_f, text="✕", width=2,
                   command=lambda: self._fname_var.set("")).pack(side="left", padx=3)

        # ファイル
        ttk.Label(self, text="ファイル：").grid(row=2, column=0, sticky="w", pady=5)
        self._fpath_var = tk.StringVar()
        ttk.Entry(self, textvariable=self._fpath_var, width=28).grid(
            row=2, column=1, sticky="ew", pady=5)
        ttk.Button(self, text="参照...", command=self._browse_file).grid(
            row=2, column=2, padx=(6, 0), pady=5)

        # シート
        ttk.Label(self, text="シート：").grid(row=3, column=0, sticky="w", pady=5)
        self._sheet_var = tk.StringVar()
        self._sheet_cb = ttk.Combobox(self, textvariable=self._sheet_var,
                                      values=[], state="readonly", width=26)
        self._sheet_cb.grid(row=3, column=1, columnspan=2, sticky="ew", pady=5)

        # 日付形式
        ttk.Label(self, text="日付形式：").grid(row=4, column=0, sticky="w", pady=5)
        dfmt_f = ttk.Frame(self)
        dfmt_f.grid(row=4, column=1, columnspan=2, sticky="w", pady=5)
        self._dfmt_var = tk.StringVar(value="M/D")
        ttk.Entry(dfmt_f, textvariable=self._dfmt_var, width=14).pack(side="left")
        ttk.Button(dfmt_f, text="⚙設定", width=6,
                   command=self._open_fmt_dialog).pack(side="left", padx=4)
        self._fmt_preview_var = tk.StringVar()
        ttk.Label(dfmt_f, textvariable=self._fmt_preview_var,
                  foreground="#0066cc").pack(side="left", padx=4)
        self._dfmt_var.trace_add("write", lambda *_: self._update_fmt_preview())
        self._update_fmt_preview()

        # 対象セル
        ttk.Label(self, text="対象セル：").grid(row=5, column=0, sticky="w", pady=5)
        tgt_f = ttk.Frame(self)
        tgt_f.grid(row=5, column=1, columnspan=2, sticky="ew", pady=5)
        self._target_var = tk.StringVar()
        ttk.Entry(tgt_f, textvariable=self._target_var, width=20).pack(
            side="left", fill="x", expand=True)
        ttk.Button(tgt_f, text="Excelで選択",
                   command=self._select_target).pack(side="left", padx=3)
        ttk.Button(tgt_f, text="✕", width=2,
                   command=lambda: self._target_var.set("")).pack(side="left")

        # 除外セル
        ttk.Label(self, text="除外セル：").grid(row=6, column=0, sticky="w", pady=5)
        exc_f = ttk.Frame(self)
        exc_f.grid(row=6, column=1, columnspan=2, sticky="ew", pady=5)
        self._exclude_var = tk.StringVar()
        ttk.Entry(exc_f, textvariable=self._exclude_var, width=20).pack(
            side="left", fill="x", expand=True)
        ttk.Button(exc_f, text="Excelで選択",
                   command=self._select_exclude).pack(side="left", padx=3)
        ttk.Button(exc_f, text="✕", width=2,
                   command=lambda: self._exclude_var.set("")).pack(side="left")

        # 日付
        ttk.Label(self, text="日付：").grid(row=7, column=0, sticky="w", pady=5)
        dt_f = ttk.Frame(self)
        dt_f.grid(row=7, column=1, columnspan=2, sticky="w", pady=5)
        self._date_var = tk.StringVar(value=date.today().isoformat())
        ttk.Entry(dt_f, textvariable=self._date_var, width=18).pack(side="left")
        ttk.Button(dt_f, text="📅", width=3,
                   command=lambda: pick_date(self, self._date_var)).pack(side="left", padx=3)

        ttk.Separator(self, orient="horizontal").grid(
            row=8, column=0, columnspan=3, sticky="ew", pady=12)

        btn_f = ttk.Frame(self)
        btn_f.grid(row=9, column=0, columnspan=3)
        ttk.Button(btn_f, text="このパターンを保存",
                   command=self._save_pattern, width=18).pack(side="left", padx=6)
        self._exec_btn = ttk.Button(btn_f, text="実 行", command=self._execute, width=12)
        self._exec_btn.pack(side="left", padx=6)

        st_lf = ttk.LabelFrame(self, text="ステータス", padding=8)
        st_lf.grid(row=10, column=0, columnspan=3, sticky="ew", pady=(14, 0))
        st_lf.columnconfigure(0, weight=1)
        self._status_var = tk.StringVar(value="")
        ttk.Label(st_lf, textvariable=self._status_var,
                  font=("Yu Gothic UI", 10), wraplength=490, justify="left").grid(
            row=0, column=0, sticky="w")

    def _open_fmt_dialog(self):
        DateFormatDialog(self, self._dfmt_var.get(),
                         lambda v: self._dfmt_var.set(v))

    def _update_fmt_preview(self):
        self._fmt_preview_var.set(f"→ {format_date_preview(self._dfmt_var.get())}")

    def _clear_all(self):
        self._fname_var.set("")
        self._fpath_var.set("")
        self._sheet_var.set("")
        self._sheet_cb["values"] = []
        self._target_var.set("")
        self._exclude_var.set("")
        self._dfmt_var.set("M/D")
        self._date_var.set(date.today().isoformat())
        self._status_var.set("")
        self._sheet_names = []

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択", initialdir=self._app.dropbox_root,
            filetypes=[("Excel ファイル", "*.xlsx *.xlsm *.xls")])
        if path:
            self._fpath_var.set(os.path.relpath(path, self._app.dropbox_root))
            self._reload_sheets()

    def _reload_sheets(self):
        full = os.path.join(self._app.dropbox_root, self._fpath_var.get())
        try:
            import openpyxl
            wb = openpyxl.load_workbook(full, read_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
            self._sheet_cb["values"] = self._sheet_names
            if self._sheet_names:
                self._sheet_cb.current(0)
        except Exception:
            self._sheet_names = []

    def _full_path(self) -> str:
        return os.path.join(self._app.dropbox_root, self._fpath_var.get())

    def _select_target(self):
        CellSelectDialog(self, "対象セルを選択", self._full_path(),
                         self._target_var.get(), lambda v: self._target_var.set(v))

    def _select_exclude(self):
        CellSelectDialog(self, "除外セルを選択", self._full_path(),
                         self._exclude_var.get(), lambda v: self._exclude_var.set(v))

    def _save_pattern(self):
        default = (self._fname_var.get().strip() or self._sheet_var.get() or "新規") + "_通常"
        dlg = tk.Toplevel(self)
        dlg.title("パターン名を入力")
        dlg.resizable(False, False)
        dlg.grab_set()
        var = tk.StringVar(value=default)
        ttk.Label(dlg, text="パターン名：").pack(padx=16, pady=(14, 4))
        ttk.Entry(dlg, textvariable=var, width=28).pack(padx=16)

        def do_save():
            name = var.get().strip()
            if not name:
                return
            self._app.store.save_pattern({
                "pattern_name" : name,
                "facility_name": self._fname_var.get().strip(),
                "file_path"    : self._fpath_var.get().strip(),
                "sheet_name"   : self._sheet_var.get().strip(),
                "date_format"  : self._dfmt_var.get().strip(),
                "target_cells" : parse_cells(self._target_var.get()),
                "exclude_cells": parse_cells(self._exclude_var.get()),
                "last_executed": "",
                "last_status"  : "",
            })
            self._app.refresh_pattern_lists()
            dlg.destroy()
            self._status_var.set(f"✓ パターン「{name}」を保存しました。")

        ttk.Button(dlg, text="保存", command=do_save, width=12).pack(pady=10)
        center_window(dlg, self)

    def load_pattern(self, pattern: dict):
        self._fname_var.set(pattern.get("facility_name", ""))
        self._fpath_var.set(pattern.get("file_path", ""))
        self._dfmt_var.set(pattern.get("date_format", "M/D"))
        self._target_var.set(cells_to_str(pattern.get("target_cells", [])))
        self._exclude_var.set(cells_to_str(pattern.get("exclude_cells", [])))
        self._reload_sheets()
        sheet = pattern.get("sheet_name", "")
        if sheet in self._sheet_names:
            self._sheet_var.set(sheet)

    def _execute(self):
        fpath   = self._fpath_var.get().strip()
        sheet   = self._sheet_var.get().strip()
        date_s  = self._date_var.get().strip()
        targets = parse_cells(self._target_var.get())

        if not fpath:   self._status_var.set("⚠ ファイルを選択してください。"); return
        if not sheet:   self._status_var.set("⚠ シートを選択してください。"); return
        if not targets: self._status_var.set("⚠ 対象セルを入力してください。"); return
        try:
            date_obj = date.fromisoformat(date_s)
        except ValueError:
            self._status_var.set("⚠ 日付の形式が正しくありません（YYYY-MM-DD）。"); return

        exclude = parse_cells(self._exclude_var.get())
        fmt     = self._dfmt_var.get()

        dlg = ConfirmDialog(self, (
            f"【{self._fname_var.get() or sheet}】の Excel に\n"
            f"日付【{format_date_preview(fmt, date_obj)}】を書き込みます。\n\n"
            f"対象セル: {cells_to_str(targets)}\n"
            f"除外セル: {cells_to_str(exclude) or 'なし'}\n\n"
            "処理します。よろしいですか？"
        ))

        if dlg.result == ConfirmDialog.RESULT_OPEN:
            cell_selector.open_file_in_excel(os.path.join(self._app.dropbox_root, fpath))
            self._status_var.set("ファイルを開きました。処理はキャンセルされました。")
            return
        if dlg.result != ConfirmDialog.RESULT_YES:
            self._status_var.set("キャンセルしました。")
            return

        self._exec_btn.config(state="disabled")
        self._status_var.set("書き込み中...")
        self.update_idletasks()

        target_infos = [{
            "file_path": fpath, "sheet_name": sheet,
            "date_format": fmt, "type": "cell", "address": c,
        } for c in targets]

        result = self._app.updater.update_dates(target_infos, date_obj, exclude)
        ok_c  = sum(1 for d in result["details"] if d["status"] == "ok")
        err_c = sum(1 for d in result["details"] if d["status"] == "error")
        lines = [f"{'✓' if result['success'] else '⚠'} 完了（成功 {ok_c} 件 / 失敗 {err_c} 件）"]
        for d in result["details"]:
            lines.append(f"  {d['message']}")
        self._status_var.set("\n".join(lines))

        self._app.logger.append(
            pattern_name=self._fname_var.get() or sheet,
            facility_name=self._fname_var.get(),
            file_path=fpath, sheet_name=sheet,
            target_cells=targets, exclude_cells=exclude,
            write_date=date_s, success=result["success"],
            detail="; ".join(d["message"] for d in result["details"]),
        )
        self._app.refresh_log()
        self._exec_btn.config(state="normal")

        if not result["success"]:
            messagebox.showerror("エラーあり",
                                 "\n".join(d["message"] for d in result["details"]
                                           if d["status"] == "error"))


# ══════════════════════════════════════════════════════
# タブ② 一括実行
# ══════════════════════════════════════════════════════

class BulkTab(ttk.Frame):
    def __init__(self, parent, app: "DateUpdateApp"):
        super().__init__(parent, padding=10)
        self._app = app
        self._items: list[dict] = []
        self._current_idx: int = 0
        self._scan_results = []
        self._scan_hits: list[dict] = []
        self._build()

    def _build(self):
        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        # ── 左ペイン
        left = ttk.Frame(self)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ttk.Label(left, text="施設リスト", font=("Yu Gothic UI", 11, "bold")).pack(anchor="w")

        lf = ttk.Frame(left)
        lf.pack(fill="both", expand=True, pady=4)
        self._listbox = tk.Listbox(lf, width=26, selectmode="single", font=("Yu Gothic UI", 10))
        sb = ttk.Scrollbar(lf, orient="vertical", command=self._listbox.yview)
        self._listbox.config(yscrollcommand=sb.set)
        self._listbox.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._listbox.bind("<<ListboxSelect>>", self._on_list_select)

        br = ttk.Frame(left)
        br.pack(fill="x", pady=4)
        ttk.Button(br, text="追加", command=self._add_item, width=8).pack(side="left", padx=2)
        ttk.Button(br, text="削除", command=self._remove_item, width=8).pack(side="left", padx=2)

        # ── 右ペイン
        right = ttk.Frame(self)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(1, weight=1)

        ttk.Label(right, text="詳細", font=("Yu Gothic UI", 11, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

        detail_keys = [("施設名：","_d_fname"),("ファイル：","_d_fpath"),
                       ("シート：","_d_sheet"),("対象セル：","_d_target"),("除外セル：","_d_exclude")]
        self._detail_vars: dict[str, tk.StringVar] = {}
        for i, (label, key) in enumerate(detail_keys, start=1):
            ttk.Label(right, text=label).grid(row=i, column=0, sticky="w", pady=3)
            var = tk.StringVar()
            self._detail_vars[key] = var
            ttk.Entry(right, textvariable=var, width=32, state="readonly").grid(
                row=i, column=1, sticky="ew", pady=3)

        self._include_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(right, text="この施設を一括実行に含める",
                        variable=self._include_var,
                        command=self._toggle_include).grid(
            row=6, column=0, columnspan=2, sticky="w", pady=6)

        nav = ttk.Frame(right)
        nav.grid(row=7, column=0, columnspan=2, sticky="ew", pady=6)
        self._prev_btn = ttk.Button(nav, text="← 戻る", command=self._prev, width=10)
        self._prev_btn.pack(side="left", padx=4)
        self._next_btn = ttk.Button(nav, text="進む →", command=self._next, width=10)
        self._next_btn.pack(side="left", padx=4)

        ttk.Separator(right, orient="horizontal").grid(
            row=8, column=0, columnspan=2, sticky="ew", pady=6)

        # スキャン
        scan_lf = ttk.LabelFrame(right, text="Dropboxスキャンで追加", padding=8)
        scan_lf.grid(row=9, column=0, columnspan=2, sticky="ew", pady=4)
        scan_lf.columnconfigure(1, weight=1)

        ttk.Label(scan_lf, text="フォルダ：").grid(row=0, column=0, sticky="w")
        self._scan_folder_var = tk.StringVar()
        self._scan_folder_cb = ttk.Combobox(scan_lf, textvariable=self._scan_folder_var, width=22)
        self._scan_folder_cb.grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(scan_lf, text="参照...", command=self._browse_scan_folder).grid(row=0, column=2)

        self._subdir_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(scan_lf, text="サブフォルダ含む",
                        variable=self._subdir_var).grid(row=1, column=0, columnspan=2, sticky="w", pady=2)

        ttk.Label(scan_lf, text="検索：").grid(row=2, column=0, sticky="w", pady=2)
        self._search_var = tk.StringVar()
        ttk.Entry(scan_lf, textvariable=self._search_var, width=16).grid(
            row=2, column=1, sticky="ew", padx=4, pady=2)
        self._search_mode = tk.StringVar(value="OR")
        ttk.Radiobutton(scan_lf, text="AND", variable=self._search_mode, value="AND").grid(row=2, column=2)
        ttk.Radiobutton(scan_lf, text="OR",  variable=self._search_mode, value="OR").grid(row=2, column=3)

        sbr = ttk.Frame(scan_lf)
        sbr.grid(row=3, column=0, columnspan=4, sticky="w", pady=4)
        ttk.Button(sbr, text="スキャン開始", command=self._scan).pack(side="left", padx=4)
        ttk.Button(sbr, text="検索",         command=self._search_scan).pack(side="left", padx=4)

        srf = ttk.Frame(scan_lf)
        srf.grid(row=4, column=0, columnspan=4, sticky="ew", pady=4)
        srf.columnconfigure(0, weight=1)
        self._scan_lb = tk.Listbox(srf, height=4, font=("Yu Gothic UI", 9))
        ssb = ttk.Scrollbar(srf, orient="vertical", command=self._scan_lb.yview)
        self._scan_lb.config(yscrollcommand=ssb.set)
        self._scan_lb.pack(side="left", fill="both", expand=True)
        ssb.pack(side="right", fill="y")

        ttk.Button(scan_lf, text="一括リストに追加",
                   command=self._add_from_scan).grid(row=5, column=0, columnspan=4, pady=4)
        self._scan_status = tk.StringVar(value="")
        ttk.Label(scan_lf, textvariable=self._scan_status, foreground="#555").grid(
            row=6, column=0, columnspan=4, sticky="w")

        # 一括実行フッター
        bulk_row = ttk.Frame(self)
        bulk_row.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        ttk.Label(bulk_row, text="一括実行日付：").pack(side="left")
        self._bulk_date_var = tk.StringVar(value=date.today().isoformat())
        ttk.Entry(bulk_row, textvariable=self._bulk_date_var, width=14).pack(side="left", padx=4)
        ttk.Button(bulk_row, text="📅", width=3,
                   command=lambda: pick_date(self, self._bulk_date_var)).pack(side="left")
        ttk.Button(bulk_row, text="一括実行", command=self._bulk_execute,
                   width=14).pack(side="right", padx=4)

        self._bulk_status = tk.StringVar(value="")
        ttk.Label(self, textvariable=self._bulk_status, foreground="#333",
                  wraplength=540, justify="left").grid(
            row=2, column=0, columnspan=2, sticky="w", pady=4)

    # ── リスト操作 ─────────────────────────────

    def refresh_list(self):
        self._app.store.reload()
        config_names = self._app.loader.get_facility_names()
        existing = {p["pattern_name"] for p in self._items}

        for cname in config_names:
            plist = self._app.store.get_by_facility(cname)
            if not plist:
                dummy = {
                    "pattern_name": f"{cname}_新規", "facility_name": cname,
                    "file_path": "", "sheet_name": "", "date_format": "M/D",
                    "target_cells": [], "exclude_cells": [],
                    "last_executed": "", "last_status": "未設定", "_checked": False,
                }
                if dummy["pattern_name"] not in existing:
                    self._items.append(dummy)
                    existing.add(dummy["pattern_name"])
            else:
                for p in plist:
                    p.setdefault("_checked", True)
                    if p["pattern_name"] not in existing:
                        self._items.append(p)
                        existing.add(p["pattern_name"])
        self._redraw_list()

    def add_patterns_from_history(self, patterns: list[dict]) -> int:
        existing = {p["pattern_name"] for p in self._items}
        added = 0
        for p in patterns:
            if p["pattern_name"] not in existing:
                pc = copy.deepcopy(p)
                pc["_checked"] = True
                self._items.append(pc)
                existing.add(pc["pattern_name"])
                added += 1
        if added:
            self._redraw_list()
        return added

    def _redraw_list(self):
        self._listbox.delete(0, "end")
        for item in self._items:
            status  = item.get("last_status", "") or "未実行"
            checked = "☑" if item.get("_checked", True) else "☐"
            label   = f"{checked} {item.get('facility_name') or item['pattern_name']} [{status}]"
            self._listbox.insert("end", label)
        if self._items:
            idx = min(self._current_idx, len(self._items) - 1)
            self._listbox.selection_set(idx)
            self._show_detail(idx)
        self._update_nav_buttons()

    def _on_list_select(self, event):
        sel = self._listbox.curselection()
        if sel:
            self._current_idx = sel[0]
            self._show_detail(self._current_idx)

    def _show_detail(self, idx: int):
        if not (0 <= idx < len(self._items)):
            return
        p = self._items[idx]
        self._detail_vars["_d_fname"].set(p.get("facility_name", ""))
        self._detail_vars["_d_fpath"].set(p.get("file_path", ""))
        self._detail_vars["_d_sheet"].set(p.get("sheet_name", ""))
        self._detail_vars["_d_target"].set(cells_to_str(p.get("target_cells", [])))
        self._detail_vars["_d_exclude"].set(cells_to_str(p.get("exclude_cells", [])))
        self._include_var.set(p.get("_checked", True))
        self._update_nav_buttons()

    def _toggle_include(self):
        if 0 <= self._current_idx < len(self._items):
            self._items[self._current_idx]["_checked"] = self._include_var.get()
            self._redraw_list()

    def _update_nav_buttons(self):
        self._prev_btn.config(state="disabled" if self._current_idx == 0 else "normal")
        self._next_btn.config(
            state="disabled" if self._current_idx >= len(self._items) - 1 else "normal")

    def _prev(self):
        if self._current_idx > 0:
            self._current_idx -= 1
            self._listbox.selection_clear(0, "end")
            self._listbox.selection_set(self._current_idx)
            self._show_detail(self._current_idx)

    def _next(self):
        if self._current_idx < len(self._items) - 1:
            self._current_idx += 1
            self._listbox.selection_clear(0, "end")
            self._listbox.selection_set(self._current_idx)
            self._show_detail(self._current_idx)

    def _add_item(self):
        dlg = tk.Toplevel(self)
        dlg.title("追加方法を選択")
        dlg.resizable(False, False)
        dlg.grab_set()
        ttk.Label(dlg, text="どちらの方法で追加しますか？",
                  font=("Yu Gothic UI", 11)).pack(padx=20, pady=(16, 10))
        bf = ttk.Frame(dlg)
        bf.pack(pady=(0, 16))

        def open_new():
            dlg.destroy()
            def on_save(pattern):
                pattern["_checked"] = True
                self._app.store.save_pattern(pattern)
                self._items.append(pattern)
                self._redraw_list()
            PatternEditDialog(self, self._app.dropbox_root, None, on_save)

        def open_select():
            dlg.destroy()
            PatternSelectDialog(self, self._app.store,
                                lambda pats: self._on_patterns_selected(pats))

        ttk.Button(bf, text="新規パターン設定から追加",
                   command=open_new, width=22).grid(row=0, column=0, padx=6, pady=4)
        ttk.Button(bf, text="保存済みパターンから選択",
                   command=open_select, width=22).grid(row=1, column=0, padx=6, pady=4)
        ttk.Button(bf, text="キャンセル",
                   command=dlg.destroy, width=12).grid(row=2, column=0, pady=4)
        center_window(dlg, self)

    def _on_patterns_selected(self, patterns: list[dict]):
        added = self.add_patterns_from_history(patterns)
        self._bulk_status.set(f"✓ {added} 件追加しました。")

    def _remove_item(self):
        if not (0 <= self._current_idx < len(self._items)):
            return
        name = self._items[self._current_idx]["pattern_name"]
        if messagebox.askyesno("確認", f"「{name}」をリストから削除しますか？"):
            self._items.pop(self._current_idx)
            self._current_idx = max(0, self._current_idx - 1)
            self._redraw_list()

    def _refresh_scan_folders(self):
        folders = self._app.scanner.get_top_folders()
        self._scan_folder_cb["values"] = folders
        if folders and not self._scan_folder_var.get():
            self._scan_folder_var.set(folders[0])

    def _browse_scan_folder(self):
        path = filedialog.askdirectory(initialdir=self._app.dropbox_root)
        if path:
            self._scan_folder_var.set(os.path.relpath(path, self._app.dropbox_root))

    def _scan(self):
        folder = self._scan_folder_var.get().strip()
        if not folder:
            self._scan_status.set("⚠ スキャンするフォルダを選択してください。"); return
        self._scan_status.set("スキャン中...")
        self.update_idletasks()

        def do_scan():
            results = self._app.scanner.scan_folder(
                folder, self._subdir_var.get(),
                progress_callback=lambda f: self._scan_status.set(f"スキャン中: {f}"))
            self._scan_results = results
            self.after(0, lambda: self._show_scan_results(results))

        threading.Thread(target=do_scan, daemon=True).start()

    def _search_scan(self):
        if not self._scan_results:
            self._scan(); return
        hits = self._app.scanner.search(
            self._scan_results, self._search_var.get(), self._search_mode.get())
        self._display_scan_hits(hits)

    def _show_scan_results(self, results):
        hits = self._app.scanner.search(
            results, self._search_var.get(), self._search_mode.get())
        self._display_scan_hits(hits)
        self._scan_status.set(f"✓ スキャン完了（{len(results)} ファイル / {len(hits)} 件表示）")

    def _display_scan_hits(self, hits):
        self._scan_lb.delete(0, "end")
        self._scan_hits = hits
        for h in hits:
            self._scan_lb.insert("end", h["label"])

    def _add_from_scan(self):
        sel = self._scan_lb.curselection()
        if not sel:
            self._scan_status.set("⚠ 追加する項目を選択してください。"); return
        hit = self._scan_hits[sel[0]]
        initial = {
            "pattern_name": f"{hit['file_name']}_{hit['sheet_name']}",
            "facility_name": hit["file_name"],
            "file_path": hit["file_path"],
            "sheet_name": hit["sheet_name"],
            "date_format": "M/D",
            "target_cells": [], "exclude_cells": [],
        }

        def on_save(pattern):
            pattern["_checked"] = True
            self._app.store.save_pattern(pattern)
            self._items.append(pattern)
            self._redraw_list()
            self._scan_status.set(f"✓「{pattern['pattern_name']}」を追加しました。")

        PatternEditDialog(self, self._app.dropbox_root, initial, on_save)

    def _bulk_execute(self):
        checked = [p for p in self._items if p.get("_checked", True)]
        if not checked:
            self._bulk_status.set("⚠ 実行対象の施設にチェックを入れてください。"); return

        unconfigured = [p for p in checked if not p.get("file_path") or not p.get("target_cells")]
        if unconfigured:
            messagebox.showwarning("設定不足",
                                   "以下のパターンにファイル・対象セルが設定されていません。\n" +
                                   ", ".join(p["pattern_name"] for p in unconfigured))
            return

        date_s = self._bulk_date_var.get().strip()
        try:
            date_obj = date.fromisoformat(date_s)
        except ValueError:
            self._bulk_status.set("⚠ 日付の形式が正しくありません（YYYY-MM-DD）。"); return

        dlg = ConfirmDialog(self, f"{len(checked)} 施設に日付【{date_s}】を書き込みます。\n処理します。よろしいですか？")
        if dlg.result != ConfirmDialog.RESULT_YES:
            self._bulk_status.set("キャンセルしました。"); return

        lines = []
        for p in checked:
            target_infos = [{
                "file_path": p["file_path"], "sheet_name": p["sheet_name"],
                "date_format": p.get("date_format", "M/D"),
                "type": "cell", "address": c,
            } for c in p["target_cells"]]

            result = self._app.updater.update_dates(
                target_infos, date_obj, p.get("exclude_cells", []))
            ok_c   = sum(1 for d in result["details"] if d["status"] == "ok")
            status = "成功" if result["success"] else "失敗"
            lines.append(
                f"{'✓' if result['success'] else '⚠'} "
                f"{p.get('facility_name') or p['pattern_name']}: {status} ({ok_c}件)")

            p["last_executed"] = date_s
            p["last_status"]   = status
            if p.get("pattern_name"):
                self._app.store.update_execution_result(p["pattern_name"], result["success"])

            self._app.logger.append(
                pattern_name=p.get("pattern_name", ""),
                facility_name=p.get("facility_name", ""),
                file_path=p["file_path"], sheet_name=p["sheet_name"],
                target_cells=p["target_cells"], exclude_cells=p.get("exclude_cells", []),
                write_date=date_s, success=result["success"],
                detail="; ".join(d["message"] for d in result["details"]),
            )

        self._bulk_status.set("\n".join(lines))
        self._redraw_list()
        self._app.refresh_log()
        self._app.refresh_pattern_lists()


# ══════════════════════════════════════════════════════
# タブ③ 履歴・パターン管理
# ══════════════════════════════════════════════════════

class HistoryTab(ttk.Frame):
    def __init__(self, parent, app: "DateUpdateApp"):
        super().__init__(parent, padding=16)
        self._app = app
        self._all_patterns: list[dict] = []
        self._all_logs: list[dict]     = []
        self._build()

    def _build(self):
        self.columnconfigure(0, weight=1)

        # ── 保存済みパターン ヘッダー
        ph = ttk.Frame(self)
        ph.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        ttk.Label(ph, text="保存済みパターン",
                  font=("Yu Gothic UI", 11, "bold")).pack(side="left")
        ttk.Label(ph, text="検索：").pack(side="left", padx=(16, 2))
        self._pat_search_var  = tk.StringVar()
        self._pat_search_mode = tk.StringVar(value="OR")
        ttk.Entry(ph, textvariable=self._pat_search_var, width=16).pack(side="left")
        ttk.Radiobutton(ph, text="AND", variable=self._pat_search_mode, value="AND").pack(side="left", padx=2)
        ttk.Radiobutton(ph, text="OR",  variable=self._pat_search_mode, value="OR").pack(side="left", padx=2)
        ttk.Button(ph, text="検索",   command=self._search_patterns, width=6).pack(side="left", padx=3)
        ttk.Button(ph, text="クリア", command=self._clear_pat_search, width=6).pack(side="left")

        # パターンツリー
        pf = ttk.Frame(self)
        pf.grid(row=1, column=0, sticky="ew")
        pf.columnconfigure(0, weight=1)
        cols = ("pattern_name","facility_name","last_executed","last_status")
        self._pat_tree = ttk.Treeview(pf, columns=cols, show="headings", height=8, selectmode="browse")
        for col, text, w in [("pattern_name","パターン名",180),("facility_name","施設名",130),
                              ("last_executed","最終実行日時",150),("last_status","ステータス",80)]:
            self._pat_tree.heading(col, text=text)
            self._pat_tree.column(col, width=w)
        pat_sb = ttk.Scrollbar(pf, orient="vertical", command=self._pat_tree.yview)
        self._pat_tree.config(yscrollcommand=pat_sb.set)
        self._pat_tree.grid(row=0, column=0, sticky="ew")
        pat_sb.grid(row=0, column=1, sticky="ns")

        # パターンボタン
        br = ttk.Frame(self)
        br.grid(row=2, column=0, sticky="w", pady=6)
        ttk.Button(br, text="単体実行タブで開く",  command=self._load_to_single, width=18).pack(side="left", padx=3)
        ttk.Button(br, text="一括実行タブへ追加",  command=self._add_to_bulk,    width=18).pack(side="left", padx=3)
        ttk.Button(br, text="編集",               command=self._edit_pattern,   width=8).pack(side="left", padx=3)
        ttk.Button(br, text="複写",               command=self._copy_pattern,   width=8).pack(side="left", padx=3)
        ttk.Button(br, text="削除",               command=self._delete_pattern, width=8).pack(side="left", padx=3)

        ttk.Separator(self, orient="horizontal").grid(row=3, column=0, sticky="ew", pady=10)

        # ── 実行ログ ヘッダー
        lh = ttk.Frame(self)
        lh.grid(row=4, column=0, sticky="ew", pady=(0, 4))
        ttk.Label(lh, text="実行ログ",
                  font=("Yu Gothic UI", 11, "bold")).pack(side="left")
        ttk.Label(lh, text="検索：").pack(side="left", padx=(16, 2))
        self._log_search_var  = tk.StringVar()
        self._log_search_mode = tk.StringVar(value="OR")
        ttk.Entry(lh, textvariable=self._log_search_var, width=16).pack(side="left")
        ttk.Radiobutton(lh, text="AND", variable=self._log_search_mode, value="AND").pack(side="left", padx=2)
        ttk.Radiobutton(lh, text="OR",  variable=self._log_search_mode, value="OR").pack(side="left", padx=2)
        ttk.Button(lh, text="検索",   command=self._search_logs,    width=6).pack(side="left", padx=3)
        ttk.Button(lh, text="クリア", command=self._clear_log_search, width=6).pack(side="left")

        # ログツリー
        lfr = ttk.Frame(self)
        lfr.grid(row=5, column=0, sticky="ew")
        lfr.columnconfigure(0, weight=1)
        log_cols = ("datetime","user","facility_name","write_date","status")
        self._log_tree = ttk.Treeview(lfr, columns=log_cols, show="headings", height=6, selectmode="browse")
        for col, text, w in [("datetime","日時",150),("user","ユーザー",100),
                              ("facility_name","施設名",150),("write_date","書込日付",100),("status","結果",70)]:
            self._log_tree.heading(col, text=text)
            self._log_tree.column(col, width=w)
        log_sb = ttk.Scrollbar(lfr, orient="vertical", command=self._log_tree.yview)
        self._log_tree.config(yscrollcommand=log_sb.set)
        self._log_tree.grid(row=0, column=0, sticky="ew")
        log_sb.grid(row=0, column=1, sticky="ns")

        ttk.Button(self, text="ログをCSVで保存", command=self._export_log,
                   width=18).grid(row=6, column=0, sticky="w", pady=6)

        ttk.Separator(self, orient="horizontal").grid(row=7, column=0, sticky="ew", pady=10)

        io_r = ttk.Frame(self)
        io_r.grid(row=8, column=0, sticky="w")
        ttk.Button(io_r, text="設定エクスポート", command=self._export_patterns, width=16).pack(side="left", padx=4)
        ttk.Button(io_r, text="設定インポート",   command=self._import_patterns, width=16).pack(side="left", padx=4)

    # ── パターン検索 ──────────────────────────

    def _search_patterns(self):
        keys = ["pattern_name","facility_name","last_executed","last_status"]
        self._render_patterns(search_rows(self._all_patterns, keys,
                                          self._pat_search_var.get(),
                                          self._pat_search_mode.get()))

    def _clear_pat_search(self):
        self._pat_search_var.set("")
        self._render_patterns(self._all_patterns)

    def _render_patterns(self, patterns: list[dict]):
        self._pat_tree.delete(*self._pat_tree.get_children())
        for p in patterns:
            self._pat_tree.insert("", "end", values=(
                p.get("pattern_name",  ""),
                p.get("facility_name", ""),
                p.get("last_executed", "") or "未実行",
                p.get("last_status",   "") or "—",
            ))

    def refresh_patterns(self):
        self._all_patterns = self._app.store.get_all()
        self._render_patterns(self._all_patterns)

    def _selected_pattern_name(self) -> str | None:
        sel = self._pat_tree.selection()
        return self._pat_tree.item(sel[0])["values"][0] if sel else None

    # ── パターンボタン ────────────────────────

    def _load_to_single(self):
        name = self._selected_pattern_name()
        if not name:
            messagebox.showinfo("選択なし", "パターンを選択してください。"); return
        p = self._app.store.get_by_name(name)
        if p:
            self._app.single_tab.load_pattern(p)
            self._app.notebook.select(0)

    def _add_to_bulk(self):
        name = self._selected_pattern_name()
        if not name:
            messagebox.showinfo("選択なし", "パターンを選択してください。"); return
        p = self._app.store.get_by_name(name)
        if p:
            added = self._app.bulk_tab.add_patterns_from_history([p])
            if added:
                messagebox.showinfo("追加完了", f"「{name}」を一括実行リストに追加しました。")
                self._app.notebook.select(1)
            else:
                messagebox.showinfo("重複", f"「{name}」はすでに一括実行リストにあります。")

    def _edit_pattern(self):
        name = self._selected_pattern_name()
        if not name:
            messagebox.showinfo("選択なし", "パターンを選択してください。"); return
        p = self._app.store.get_by_name(name)

        def on_save(pattern):
            self._app.store.save_pattern(pattern)
            self._app.refresh_pattern_lists()

        PatternEditDialog(self, self._app.dropbox_root, p, on_save, allow_rename_save=True)

    def _copy_pattern(self):
        name = self._selected_pattern_name()
        if not name:
            messagebox.showinfo("選択なし", "パターンを選択してください。"); return
        p = self._app.store.get_by_name(name)
        if not p:
            return
        copied = copy.deepcopy(p)
        copied["pattern_name"]  = name + "_コピー"
        copied["last_executed"] = ""
        copied["last_status"]   = ""

        def on_save(pattern):
            self._app.store.save_pattern(pattern)
            self._app.refresh_pattern_lists()

        PatternEditDialog(self, self._app.dropbox_root, copied, on_save, allow_rename_save=True)

    def _delete_pattern(self):
        name = self._selected_pattern_name()
        if not name:
            return
        if messagebox.askyesno("確認", f"「{name}」を削除しますか？"):
            self._app.store.delete_pattern(name)
            self._app.refresh_pattern_lists()

    # ── ログ検索 ─────────────────────────────

    def _search_logs(self):
        keys = ["datetime","user","facility_name","write_date","status"]
        self._render_logs(search_rows(self._all_logs, keys,
                                      self._log_search_var.get(),
                                      self._log_search_mode.get()))

    def _clear_log_search(self):
        self._log_search_var.set("")
        self._render_logs(self._all_logs)

    def _render_logs(self, logs: list[dict]):
        self._log_tree.delete(*self._log_tree.get_children())
        for row in reversed(logs):
            self._log_tree.insert("", "end", values=(
                row.get("datetime",""), row.get("user",""),
                row.get("facility_name",""), row.get("write_date",""), row.get("status",""),
            ))

    def refresh_log(self):
        self._all_logs = self._app.logger.read_all()
        self._render_logs(self._all_logs)

    def _export_log(self):
        path = filedialog.asksaveasfilename(
            title="ログの保存先", defaultextension=".csv", filetypes=[("CSV","*.csv")])
        if path:
            ok = self._app.logger.export_to(path)
            messagebox.showinfo("完了" if ok else "エラー",
                                "保存しました。" if ok else "保存に失敗しました。")

    def _export_patterns(self):
        path = filedialog.asksaveasfilename(
            title="パターンのエクスポート先", defaultextension=".json", filetypes=[("JSON","*.json")])
        if path:
            ok = self._app.store.export_to(path)
            messagebox.showinfo("完了" if ok else "エラー",
                                "エクスポートしました。" if ok else "エクスポートに失敗しました。")

    def _import_patterns(self):
        path = filedialog.askopenfilename(
            title="インポートするJSONを選択", filetypes=[("JSON","*.json")])
        if not path:
            return
        overwrite = messagebox.askyesno("上書き確認", "同名パターンを上書きしますか？")
        added, skipped = self._app.store.import_from(path, overwrite)
        self._app.refresh_pattern_lists()
        messagebox.showinfo("インポート完了", f"追加: {added} 件 / スキップ: {skipped} 件")


# ══════════════════════════════════════════════════════
# メインアプリケーション
# ══════════════════════════════════════════════════════

class DateUpdateApp(tk.Tk):
    TITLE = "Dropbox Excel 日付更新ツール"
    SIZE  = "820x700"

    def __init__(self):
        super().__init__()
        self.title(self.TITLE)
        self.geometry(self.SIZE)
        self.resizable(True, True)

        self.dropbox_root = get_dropbox_path()
        self.loader  = ConfigLoader(DEFAULT_JSON_PATH)
        self.store   = PatternStore()
        self.logger  = LogManager()
        self.updater = ExcelUpdater(self.dropbox_root)
        self.scanner = DropboxScanner(self.dropbox_root)

        self._build_ui()
        self._init_data()

        if not validate_dropbox_path(self.dropbox_root):
            messagebox.showwarning(
                "Dropboxフォルダが見つかりません",
                f"以下のパスが存在しません。\n{self.dropbox_root}\n\n"
                "Dropboxがインストールされているか確認してください。",
            )

    def _build_ui(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        self.single_tab  = SingleTab(self.notebook, self)
        self.bulk_tab    = BulkTab(self.notebook, self)
        self.history_tab = HistoryTab(self.notebook, self)
        self.notebook.add(self.single_tab,  text="① 単体実行")
        self.notebook.add(self.bulk_tab,    text="② 一括実行")
        self.notebook.add(self.history_tab, text="③ 履歴・パターン管理")

    def _init_data(self):
        self.loader.load()
        self.refresh_pattern_lists()
        self.refresh_log()
        self.bulk_tab._refresh_scan_folders()

    def refresh_pattern_lists(self):
        self.store.reload()
        self.history_tab.refresh_patterns()
        self.bulk_tab.refresh_list()

    def refresh_log(self):
        self.history_tab.refresh_log()


if __name__ == "__main__":
    app = DateUpdateApp()
    app.mainloop()
