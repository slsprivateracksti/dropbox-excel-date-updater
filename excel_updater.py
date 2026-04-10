"""
excel_updater.py
openpyxl を使って Excel ファイルの指定セル／列に日付を書き込む。
"""

import os
from datetime import date
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

TargetInfo = dict[str, Any]
UpdateResult = dict[str, Any]


def _to_excel_number_format(fmt: str) -> str:
    result = fmt.replace("YYYY", "yyyy").replace("DD", "dd")
    for literal in ["年", "月", "日", ".", " "]:
        if literal in result:
            result = result.replace(literal, f'"{literal}"')
    return result


def _format_date_as_string(d: date, fmt: str) -> str:
    return (
        fmt.replace("YYYY", f"{d.year:04d}")
           .replace("MM",   f"{d.month:02d}")
           .replace("DD",   f"{d.day:02d}")
    )


class ExcelUpdater:
    def __init__(self, dropbox_root: str) -> None:
        self.dropbox_root = dropbox_root

    def update_dates(
        self,
        targets: list[TargetInfo],
        date_obj: date,
        exclude_cells: list[str] | None = None,
    ) -> UpdateResult:
        """
        targets リストの各セル／列に date_obj を書き込む。
        exclude_cells: 除外セル番地リスト（例: ["D5", "C3"]）
        """
        exclude_set = set(c.upper().strip() for c in (exclude_cells or []))
        details: list[dict] = []
        has_error = False

        grouped: dict[str, list[TargetInfo]] = {}
        for t in targets:
            grouped.setdefault(t["file_path"], []).append(t)

        for rel_path, group in grouped.items():
            full_path = os.path.join(self.dropbox_root, rel_path)
            try:
                wb = openpyxl.load_workbook(full_path)
            except FileNotFoundError:
                msg = f"ファイルが見つかりません: {full_path}"
                for t in group:
                    details.append({"target": t, "status": "error", "message": msg})
                has_error = True
                continue
            except Exception as e:
                msg = f"ファイルを開けません ({os.path.basename(full_path)}): {e}"
                for t in group:
                    details.append({"target": t, "status": "error", "message": msg})
                has_error = True
                continue

            file_dirty = False

            for target in group:
                result = self._write_target(wb, target, date_obj, exclude_set)
                details.append(result)
                if result["status"] == "ok":
                    file_dirty = True
                else:
                    has_error = True

            if file_dirty:
                try:
                    wb.save(full_path)
                except Exception as e:
                    msg = f"ファイル保存失敗 ({os.path.basename(full_path)}): {e}"
                    for detail in details:
                        if detail["target"].get("file_path") == rel_path:
                            detail["status"]  = "error"
                            detail["message"] = msg
                    has_error = True

            wb.close()

        return {"success": not has_error, "details": details}

    def _write_target(
        self,
        wb: openpyxl.Workbook,
        target: TargetInfo,
        date_obj: date,
        exclude_set: set[str],
    ) -> dict:
        sheet_name  = target.get("sheet_name", "")
        date_format = target.get("date_format", "YYYY/MM/DD")
        target_type = target.get("type", "cell")

        try:
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"シートが見つかりません: '{sheet_name}'")
            ws = wb[sheet_name]
        except Exception as e:
            return {"target": target, "status": "error", "message": str(e)}

        number_fmt = _to_excel_number_format(date_format)

        try:
            if target_type == "cell":
                addr = target["address"].upper()
                if addr in exclude_set:
                    return {
                        "target": target,
                        "status": "ok",
                        "message": f"⚡ {sheet_name}!{addr} → スキップ（除外セル）",
                    }
                self._write_cell(ws, addr, date_obj, number_fmt)
                msg = f"✓ {sheet_name}!{addr} → {_format_date_as_string(date_obj, date_format)}"

            elif target_type == "column":
                col     = target.get("col", "A")
                start   = int(target.get("start_row") or 1)
                end     = int(target.get("end_row")   or start)
                col_idx = column_index_from_string(col)
                written = []
                skipped = []
                for row in range(start, end + 1):
                    addr = f"{col}{row}".upper()
                    if addr in exclude_set:
                        skipped.append(addr)
                        continue
                    self._write_cell(ws, (row, col_idx), date_obj, number_fmt)
                    written.append(addr)
                msg = (
                    f"✓ {sheet_name}!{col}{start}:{col}{end} → "
                    f"{_format_date_as_string(date_obj, date_format)}"
                )
                if skipped:
                    msg += f"（除外: {', '.join(skipped)}）"
            else:
                raise ValueError(f"未対応の target type: '{target_type}'")

        except Exception as e:
            return {"target": target, "status": "error", "message": f"書き込み失敗 ({sheet_name}): {e}"}

        return {"target": target, "status": "ok", "message": msg}

    @staticmethod
    def _write_cell(ws, ref, date_obj: date, number_fmt: str) -> None:
        if isinstance(ref, tuple):
            cell = ws.cell(row=ref[0], column=ref[1])
        else:
            cell = ws[ref]
        cell.value         = date_obj
        cell.number_format = number_fmt
